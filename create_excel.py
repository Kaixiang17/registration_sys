import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 初始人員資料
participants = [
    {
        'id': 'P001',
        'name': '王小明',
        'phone': '0912345678',
        'company': '科技創新公司',
        'email': 'wang.xiaoming@tech.com',
        'qrCode': 'P001_TECH_2024',
        'registeredAt': '2024-04-01',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P002',
        'name': '李美麗',
        'phone': '0923456789',
        'company': '數位行銷集團',
        'email': 'li.meili@marketing.com',
        'qrCode': 'P002_MARKETING_2024',
        'registeredAt': '2024-04-02',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P003',
        'name': '陳建宏',
        'phone': '0934567890',
        'company': '軟體開發有限公司',
        'email': 'chen.jianhong@software.com',
        'qrCode': 'P003_SOFTWARE_2024',
        'registeredAt': '2024-04-03',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P004',
        'name': '張家豪',
        'phone': '0945678901',
        'company': '金融服務集團',
        'email': 'zhang.jiahao@finance.com',
        'qrCode': 'P004_FINANCE_2024',
        'registeredAt': '2024-04-04',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P005',
        'name': '黃芷芸',
        'phone': '0956789012',
        'company': '創意設計工作室',
        'email': 'huang.zhiyun@design.com',
        'qrCode': 'P005_DESIGN_2024',
        'registeredAt': '2024-04-05',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P006',
        'name': '王小明',
        'phone': '0967890123',
        'company': '房地產開發公司',
        'email': 'wang.xiaoming.real@realestate.com',
        'qrCode': 'P006_REALESTATE_2024',
        'registeredAt': '2024-04-06',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P007',
        'name': '王小明',
        'phone': '0978901234',
        'company': '教育培訓中心',
        'email': 'wang.xiaoming.edu@education.com',
        'qrCode': 'P007_EDUCATION_2024',
        'registeredAt': '2024-04-07',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P008',
        'name': '李美麗',
        'phone': '0989012345',
        'company': '醫療健康集團',
        'email': 'li.meili.health@healthcare.com',
        'qrCode': 'P008_HEALTHCARE_2024',
        'registeredAt': '2024-04-08',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P009',
        'name': '劉德華',
        'phone': '0990123456',
        'company': '娛樂傳媒公司',
        'email': 'liu.dehua@media.com',
        'qrCode': 'P009_MEDIA_2024',
        'registeredAt': '2024-04-09',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P010',
        'name': '周杰倫',
        'phone': '0901234567',
        'company': '音樂製作工作室',
        'email': 'zhou.jielun@music.com',
        'qrCode': 'P010_MUSIC_2024',
        'registeredAt': '2024-04-10',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P011',
        'name': '陳思宇',
        'phone': '0912345678',
        'company': '科技創新公司',
        'email': 'chen.siyu@tech.com',
        'qrCode': 'P011_TECH_2024',
        'registeredAt': '2024-04-11',
        'checkedInAt': '',
        'status': 'registered'
    },
    {
        'id': 'P012',
        'name': '黃建志',
        'phone': '0912345678',
        'company': '科技創新公司',
        'email': 'huang.jianzhi@tech.com',
        'qrCode': 'P012_TECH_2024',
        'registeredAt': '2024-04-12',
        'checkedInAt': '',
        'status': 'registered'
    }
]

# 建立 Excel 工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '人員資料'

# 定義樣式
header_font = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=12)
header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(name='微軟正黑體', size=11)
data_alignment = Alignment(horizontal='left', vertical='center')

checked_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# 欄位標題
headers = ['編號', '姓名', '電話', '公司', '電子郵件', 'QR Code', '報名日期', '報到時間', '狀態']
columns = ['id', 'name', 'phone', 'company', 'email', 'qrCode', 'registeredAt', 'checkedInAt', 'status']

# 設定欄寬
col_widths = [10, 15, 15, 25, 35, 25, 15, 25, 12]

for i, (header, width) in enumerate(zip(headers, col_widths), 1):
    col_letter = get_column_letter(i)
    ws.column_dimensions[col_letter].width = width
    cell = ws.cell(row=1, column=i, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 設定標題列高度
ws.row_dimensions[1].height = 30

# 填入資料
for row_idx, p in enumerate(participants, 2):
    ws.row_dimensions[row_idx].height = 22
    for col_idx, col_key in enumerate(columns, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=p[col_key])
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        if p['status'] == 'checked_in':
            cell.fill = checked_fill

# 凍結首列
ws.freeze_panes = 'A2'

# 儲存
wb.save('/home/ubuntu/活動報到系統/participants.xlsx')
print('Excel 檔案已建立：participants.xlsx')
